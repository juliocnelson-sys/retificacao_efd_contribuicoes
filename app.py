"""
EFD Contribuições – Retificação
Aplicativo Streamlit para uso corporativo
"""

import io
import re
import zipfile
from collections import defaultdict
from pathlib import Path

import openpyxl
import streamlit as st

# ─────────────────────────────────────────────────────────────────────────────
# Configuração da página
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="EFD Contribuições – Retificação",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown("""
<style>
    .main .block-container { padding-top: 2rem; max-width: 960px; }
    .stAlert { border-radius: 8px; }
    div[data-testid="stMetric"] { background: #f8f9fa; padding: 1rem; border-radius: 8px; border: 1px solid #e9ecef; }
    .status-ok   { color: #0F6E56; font-weight: 500; }
    .status-info { color: #185FA5; }
    .status-warn { color: #854F0B; font-weight: 500; }
    .status-err  { color: #A32D2D; font-weight: 500; }
    .log-box { background: #f8f9fa; border: 1px solid #e9ecef; border-radius: 8px;
               padding: 1rem; font-family: monospace; font-size: 13px;
               max-height: 340px; overflow-y: auto; line-height: 1.8; }
    .rule-box { background: #f8f9fa; border: 1px solid #e9ecef; border-radius: 8px; padding: 0.75rem 1rem; margin-bottom: 8px; }
    .rule-title { font-weight: 600; font-size: 14px; margin-bottom: 4px; }
    .rule-desc  { font-size: 13px; color: #555; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# Funções auxiliares de formato SPED
# ─────────────────────────────────────────────────────────────────────────────
ENCODINGS_TO_TRY = ["latin-1", "windows-1252", "utf-8", "utf-8-sig"]

def detect_encoding(data: bytes) -> str:
    """Tenta decodificar com múltiplos encodings e retorna o primeiro que funcionar."""
    for enc in ENCODINGS_TO_TRY:
        try:
            data.decode(enc)
            return enc
        except (UnicodeDecodeError, LookupError):
            continue
    return "latin-1"  # fallback seguro

def parse_pipe(line: str) -> list:
    return line.strip().strip("|").split("|")

def to_pipe(fields: list) -> str:
    return "|" + "|".join(str(f) for f in fields) + "|\n"

def fmt_valor(v) -> str:
    if v is None or str(v).strip() == "":
        return "0,00"
    s = str(v).strip().replace(".", ",")
    if "," not in s:
        s += ",00"
    return s

def soma_valores(*vals) -> float:
    """
    Soma valores numericos vindos do Excel ou de strings BR.
    Aceita: float/int do Python, strings BR ('28.616,29' ou '28616,29')
    
    REGRA: se o valor ja eh float ou int, usa diretamente.
    Se eh string, normaliza removendo pontos de milhar (antes da virgula)
    e convertendo virgula decimal para ponto.
    """
    total = 0.0
    for v in vals:
        if v is None:
            continue
        # Tipos numericos: usa diretamente, sem converter para string
        if isinstance(v, (int, float)):
            total += float(v)
            continue
        s = str(v).strip()
        if not s or s in ('← script', 'None'):
            continue
        # String com virgula E ponto: formato BR com milhar ex '28.616,29'
        if ',' in s and '.' in s:
            if s.rfind(',') > s.rfind('.'):
                # virgula = decimal, pontos = milhar -> remove pontos
                s = s.replace('.', '').replace(',', '.')
            else:
                # ponto = decimal, virgulas = milhar -> remove virgulas
                s = s.replace(',', '')
        elif ',' in s:
            # virgula decimal BR: '28616,29' -> '28616.29'
            s = s.replace(',', '.')
        # ponto decimal EN: '28616.29' -> usa como esta
        try:
            total += float(s)
        except ValueError:
            pass
    return total

def fmt_br(f: float) -> str:
    """
    Formata float para o padrao SPED: virgula decimal, sem ponto de milhar.
    Remove zeros decimais desnecessarios conforme padrao do TXT:
      148394.4  -> '148394,4'   (nao '148394,40')
      7122.93   -> '7122,93'
      0.0       -> '0'          (nao '0,00')
      2448.51   -> '2448,51'
    """
    if f == int(f):
        return str(int(f))
    # Remove zeros a direita mas mantem pelo menos 1 casa decimal
    s = f"{f:.10f}".rstrip('0')
    if s.endswith('.'):
        s += '0'
    return s.replace('.', ',')

def fmt_data(v) -> str:
    """Converte data para o formato SPED: DDMMAAAA.
    Aceita: datetime, date, string dd/mm/aaaa, string aaaa-mm-dd,
            string dd/mm/aaaa hh:mm:ss, etc.
    """
    if v is None or str(v).strip() in ("", "None"):
        return ""
    import datetime as _dt
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.strftime("%d%m%Y")
    s = str(v).strip()
    # Formato Python/ISO: 2021-10-05 00:00:00 ou 2021-10-05
    if len(s) >= 10 and s[4] == "-":
        try:
            d = _dt.datetime.strptime(s[:10], "%Y-%m-%d")
            return d.strftime("%d%m%Y")
        except ValueError:
            pass
    # Formato dd/mm/aaaa
    if len(s) >= 10 and s[2] == "/" and s[5] == "/":
        try:
            d = _dt.datetime.strptime(s[:10], "%d/%m/%Y")
            return d.strftime("%d%m%Y")
        except ValueError:
            pass
    # Formato ddmmaaaa (já correto)
    if len(s) == 8 and s.isdigit():
        return s
    return s  # retorna como está se não reconhecido

def fmt_valor_sped(v) -> str:
    """
    Formata valor numerico para o padrao SPED:
      - Decimal com virgula
      - SEM ponto separador de milhar
      - Sempre 2 casas decimais
    
    Aceita: float/int do Excel, strings BR ('1.234,56'), strings EN ('1234.56')
    Retorna: '1234,56'
    """
    if v is None:
        return "0,00"
    
    # Tipos numericos do Python/Excel: int ou float -> usa fmt_br (remove zeros)
    if isinstance(v, (int, float)):
        return fmt_br(float(v))
    
    s = str(v).strip()
    
    if s in ("", "None", "← script"):
        return "0"
    
    # String com virgula E ponto -> formato BR com milhar: '1.234,56'
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            # virgula = decimal, pontos = milhar -> remove pontos
            s = s.replace(".", "").replace(",", ".")
        else:
            # ponto = decimal, virgulas = milhar -> remove virgulas
            s = s.replace(",", "")
        try:
            return fmt_br(float(s))
        except ValueError:
            return "0"
    
    # String apenas com virgula -> decimal BR: '1234,56'
    if "," in s:
        try:
            return fmt_br(float(s.replace(',', '.')))
        except ValueError:
            return "0"
    
    # String apenas com ponto -> decimal EN: '1234.56'
    if "." in s:
        try:
            return fmt_br(float(s))
        except ValueError:
            return "0"
    
    # String sem separador: '1234'
    try:
        return fmt_br(float(s))
    except ValueError:
        return "0"

def gc(row: dict, *keys) -> str:
    for k in keys:
        if k in row and str(row[k]).strip() not in ("", "None"):
            return str(row[k]).strip()
    return ""


# ─────────────────────────────────────────────────────────────────────────────
# Leitura da planilha
# ─────────────────────────────────────────────────────────────────────────────
def sheet_to_dicts(ws) -> list:
    """
    Le uma aba do Excel e retorna lista de dicts {header: valor}.

    Logica de deteccao de cabecalho:
    - Abas simples (0150, F100...): linha 1 = titulo, linha 2 = cabecalho, linha 3+ = dados
    - Abas com grupos coloridos (A100, C100, D100, C500...):
        linha 1 = titulo, linha 2 = grupos coloridos, linha 3 = cabecalho real, linha 4+ = dados

    Para as abas com grupos, a linha 2 tem textos como "IDENTIFICACAO",
    "A100 - CABECALHO", etc., e a linha 3 tem os nomes reais dos campos
    como CNPJ_ESTAB, NUM_DOC, VL_ITEM etc.

    Estrategia: usa a linha de cabecalho como aquela que contiver CNPJ_ESTAB
    ou o maior numero de campos reconheciveis do layout SPED.
    Ignora linhas que sejam apenas descricoes de grupo (sem campos SPED).
    """
    CAMPOS_SPED = {
        'CNPJ_ESTAB', 'IND_OPER', 'IND_EMIT', 'COD_PART', 'COD_SIT',
        'NUM_DOC', 'DT_DOC', 'VL_DOC', 'IND_PGTO', 'CST_PIS',
        'CST_COFINS', 'COD_ITEM', 'VL_ITEM', 'NUM_ITEM', 'UNID',
        'COD_NAT_REC', 'COD_CTA', 'DT_ALT', 'NAT_BC_CRED', 'COD_MOD',
        'VL_BC_PIS', 'VL_PIS', 'VL_BC_COFINS', 'VL_COFINS', 'ALIQ_PIS',
        'COD_PART\n(Fornecedor/Cliente)', 'COD_PART\n(Transportadora)',
        'PERÍODO\n(MMAAAA)',
    }

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return []

    # Conta quantos campos SPED cada linha contem
    # A linha com maior contagem e o cabecalho real
    best_idx = None
    best_score = 0

    for i, row in enumerate(all_rows):
        score = 0
        for v in row:
            if v is None:
                continue
            # Pega o texto da celula (primeira linha se tiver \n)
            s = str(v).strip()
            # Verifica se e um campo SPED conhecido
            if s in CAMPOS_SPED:
                score += 2  # match exato vale mais
            elif s.split('\n')[0].strip() in CAMPOS_SPED:
                score += 2
            elif any(campo in s for campo in CAMPOS_SPED):
                score += 1
        if score > best_score:
            best_score = score
            best_idx = i

    if best_idx is None or best_score == 0:
        return []

    # Monta dicionario de cabecalhos usando o texto completo da celula
    headers = []
    for v in all_rows[best_idx]:
        if v is None:
            headers.append(None)
        else:
            headers.append(str(v).strip())

    rows = []
    for row in all_rows[best_idx + 1:]:
        # Pula linhas completamente vazias
        if all(c is None for c in row):
            continue
        # Pula linhas que sao apenas linha de exemplo com "← script"
        non_empty = [c for c in row if c is not None and str(c).strip() != '']
        if all(str(c).strip() == '← script' for c in non_empty):
            continue
        # Monta dict somente para colunas com cabecalho valido
        d = {}
        for j, (h, v) in enumerate(zip(headers, row)):
            if h is not None:
                d[h] = v if v is not None else ''
        if d:
            rows.append(d)
    return rows

@st.cache_data(show_spinner=False)
def load_planilha(file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    data = {}
    for name in wb.sheetnames:
        if name in ("INSTRUÇÕES", "LISTAS"):
            continue
        data[name] = sheet_to_dicts(wb[name])
    return data


# ─────────────────────────────────────────────────────────────────────────────
# Parse do TXT
# ─────────────────────────────────────────────────────────────────────────────
def parse_txt(lines: list) -> dict:
    result = {"cnpj_0140": {}, "cod_0200": set(), "bloco_0_end": None}
    for idx, line in enumerate(lines):
        ls = line.strip()
        if not ls.startswith("|"):
            continue
        f = parse_pipe(ls)
        reg = f[0]
        if reg == "0140":
            cnpj = f[3] if len(f) > 3 else ""
            result["cnpj_0140"][cnpj] = f
        elif reg == "0200":
            result["cod_0200"].add(f[1] if len(f) > 1 else "")
        elif reg == "0990":
            result["bloco_0_end"] = idx
    return result

def get_periodo(lines: list) -> str | None:
    for line in lines:
        if line.strip().startswith("|0000|"):
            f = parse_pipe(line.strip())
            if len(f) > 5 and f[5]:
                return f[5][2:]  # MMAAAA
    return None

def filter_periodo(rows: list, periodo: str | None) -> list:
    """
    Filtra linhas pelo período do TXT.
    Aceita período da planilha nos formatos:
      - "102021" (MMAAAA string) — match direto
      - datetime(2021,10,1)      — converte para MMAAAA
      - "01/10/2021"             — converte para MMAAAA
      - "2021-10-01"             — converte para MMAAAA
      - vazio/None               — passa tudo
    """
    import datetime as _dt
    if not periodo:
        return rows

    def normalizar_periodo(v) -> str:
        """Converte qualquer representação de período para MMAAAA."""
        if v is None or str(v).strip() == "":
            return ""
        # datetime ou date do Python/Excel
        if isinstance(v, (_dt.datetime, _dt.date)):
            return f"{v.month:02d}{v.year}"
        s = str(v).strip()
        # Já está no formato MMAAAA (6 dígitos)
        if len(s) == 6 and s.isdigit():
            return s
        # Formato ISO ou similar: "2021-10-01" ou "2021-10-01 00:00:00"
        if len(s) >= 10 and s[4] == "-":
            try:
                d = _dt.datetime.strptime(s[:10], "%Y-%m-%d")
                return f"{d.month:02d}{d.year}"
            except ValueError:
                pass
        # Formato dd/mm/aaaa
        if len(s) >= 10 and s[2] == "/" and s[5] == "/":
            try:
                d = _dt.datetime.strptime(s[:10], "%d/%m/%Y")
                return f"{d.month:02d}{d.year}"
            except ValueError:
                pass
        # Formato MM/AAAA (ex: "10/2021") — sem dia
        if len(s) == 7 and s[2] == "/":
            try:
                partes = s.split("/")
                return f"{int(partes[0]):02d}{partes[1]}"
            except (ValueError, IndexError):
                pass
        # Formato MM-AAAA (ex: "10-2021") — sem dia
        if len(s) == 7 and s[2] == "-":
            try:
                partes = s.split("-")
                return f"{int(partes[0]):02d}{partes[1]}"
            except (ValueError, IndexError):
                pass
        # Formato M/AAAA (ex: "1/2021") — mes sem zero
        if len(s) == 6 and "/" in s:
            try:
                partes = s.split("/")
                return f"{int(partes[0]):02d}{partes[1]}"
            except (ValueError, IndexError):
                pass
        return s

    result = []
    for r in rows:
        raw = r.get("PERÍODO\n(MMAAAA)", r.get("PERÍODO", r.get("PERIODO", "")))
        p = normalizar_periodo(raw)
        if not p or p == periodo:
            result.append(r)
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Construtores de registros
# ─────────────────────────────────────────────────────────────────────────────
def build_0140(row):
    cnpj = gc(row, "CNPJ_ESTAB", "CNPJ")
    return to_pipe(["0140", cnpj, gc(row, "NOME / RAZÃO SOCIAL"), cnpj, gc(row, "UF"), "", gc(row, "COD_MUN"), "", ""])

def build_0150(r):
    return to_pipe(["0150", gc(r,"COD_PART"), gc(r,"NOME"), gc(r,"COD_PAIS") or "1058",
                    gc(r,"CNPJ"), gc(r,"CPF"), gc(r,"IE"), gc(r,"COD_MUN"),
                    gc(r,"SUFRAMA"), gc(r,"END"), gc(r,"NUM"), gc(r,"COMPL"), gc(r,"BAIRRO")])

def build_0190(r):
    return to_pipe(["0190", gc(r,"UNID"), gc(r,"DESCR")])

def build_0200(r):
    return to_pipe(["0200", gc(r,"COD_ITEM"), gc(r,"DESCR_ITEM"), gc(r,"COD_BARRA"), "",
                    gc(r,"UNID_INV"), gc(r,"TIPO_ITEM"), gc(r,"COD_NCM"),
                    gc(r,"EX_IPI"), gc(r,"COD_GEN"), gc(r,"COD_LST"), fmt_valor(r.get("ALIQ_ICMS",""))])

def build_0400(r):
    return to_pipe(["0400", gc(r,"COD_NAT_REC"), gc(r,"DESCR_NAT_REC")])

def build_0500(r):
    # Layout 0500 conforme Manual EFD Contribuicoes v1.35 (9 campos):
    # REG|DT_ALT|COD_NAT_CC|IND_CTA|NIVEL|COD_CTA|NOME_CTA|COD_CTA_SUP|COD_CTA_REF
    # COD_NAT_CC: 01=Ativo 02=Passivo 03=PL 04=Resultado 05=Compensacao 09=Outras
    # Aceita tanto o nome simples ('DT_ALT') quanto o nome com descricao da planilha
    # ('DT_ALT\n(DDMMAAAA)') pois a planilha modelo v4 usa nomes compostos
    def get_campo(row, *bases):
        """Busca campo pelo nome base ou por qualquer chave que comece com esse nome."""
        for base in bases:
            # Tenta match exato primeiro
            if base in row and str(row[base]).strip() not in ('', 'None'):
                return str(row[base]).strip()
            # Tenta match por prefixo (ex: 'DT_ALT' bate em 'DT_ALT\n(DDMMAAAA)')
            for k, v in row.items():
                if str(k).split('\n')[0].strip() == base:
                    if str(v).strip() not in ('', 'None'):
                        return str(v).strip()
        return ''

    def get_data_campo(row, *bases):
        """Busca campo de data aceitando nome simples ou composto."""
        for base in bases:
            v = row.get(base, '')
            if v != '' and v is not None:
                return fmt_data(v)
            for k, val in row.items():
                if str(k).split('\n')[0].strip() == base:
                    if val not in ('', None):
                        return fmt_data(val)
        return ''

    def get_nivel(row):
        for k, v in row.items():
            if str(k).split('\n')[0].strip() in ('NÍVEL', 'NIVEL'):
                return str(v).strip() if v not in ('', None) else ''
        return str(row.get('NÍVEL', row.get('NIVEL', ''))).strip()

    return to_pipe(["0500",
        get_data_campo(r, "DT_ALT"),
        get_campo(r, "COD_NAT_CC", "COD_NAT"),
        get_campo(r, "IND_CTA"),
        get_nivel(r),
        get_campo(r, "COD_CTA"),
        get_campo(r, "NOME_CTA"),
        get_campo(r, "COD_CTA_SUP"),
        get_campo(r, "COD_CTA_REF"),
    ])

def build_a_block(grupo: list) -> list:
    # Layout A100 do Manual EFD Contribuicoes: 21 campos
    # |A100|IND_OPER|IND_EMIT|COD_PART|COD_MOD|COD_SIT|SER|SUB|NUM_DOC|CHV_DOC|
    # |DT_DOC|DT_EXE_SERV|VL_DOC|IND_PGTO|VL_DESC|VL_BC_PIS|ALIQ_PIS|VL_PIS|
    # |VL_BC_COFINS|ALIQ_COFINS|VL_COFINS|COD_CTA|COD_MUN|
    r0 = grupo[0]
    vl_doc    = soma_valores(*[r.get("VL_ITEM", 0) for r in grupo])
    vl_desc   = soma_valores(*[r.get("VL_DESC_ITEM", 0) for r in grupo])
    vl_bc_pis = soma_valores(*[r.get("VL_BC_PIS_ITEM", 0) for r in grupo])
    vl_pis    = soma_valores(*[r.get("VL_PIS_ITEM", 0) for r in grupo])
    vl_bc_cof = soma_valores(*[r.get("VL_BC_COFINS_ITEM", 0) for r in grupo])
    vl_cof    = soma_valores(*[r.get("VL_COFINS_ITEM", 0) for r in grupo])
    aliq_pis  = fmt_valor_sped(r0.get("ALIQ_PIS_ITEM", ""))
    aliq_cof  = fmt_valor_sped(r0.get("ALIQ_COFINS_ITEM", ""))
    # Layout A100: 21 campos conforme Manual EFD Contribuicoes
    # REG|IND_OPER|IND_EMIT|COD_PART|COD_MOD|COD_SIT|SER|SUB|NUM_DOC|CHV_DOC|
    # DT_DOC|DT_EXE_SERV|VL_DOC|IND_PGTO|VL_DESC|VL_BC_PIS|ALIQ_PIS|VL_PIS|
    # VL_BC_COFINS|ALIQ_COFINS|VL_COFINS
    # Layout A100 conforme Manual EFD Contribuicoes v1.35:
    # REG|IND_OPER|IND_EMIT|COD_PART|COD_SIT|SER|SUB|NUM_DOC|CHV_NFSE|
    # DT_DOC|DT_EXE_SERV|VL_DOC|IND_PGTO|VL_DESC|VL_BC_PIS|VL_PIS|
    # VL_BC_COFINS|VL_COFINS|VL_PIS_RET|VL_COFINS_RET|VL_ISS
    # NOTA: A100 NAO tem COD_MOD nem ALIQ_PIS nem ALIQ_COFINS
    lines = [to_pipe(["A100",
        gc(r0,"IND_OPER"), gc(r0,"IND_EMIT"), gc(r0,"COD_PART\n(Fornecedor/Cliente)"),
        gc(r0,"COD_SIT"), gc(r0,"SER"), gc(r0,"SUB"),
        gc(r0,"NUM_DOC"), gc(r0,"CHV_DOC"),
        fmt_data(r0.get("DT_DOC","")),
        fmt_data(r0.get("DT_EXE_SERV","")),
        fmt_br(vl_doc), gc(r0,"IND_PGTO"), fmt_br(vl_desc),
        fmt_br(vl_bc_pis), fmt_br(vl_pis),
        fmt_br(vl_bc_cof), fmt_br(vl_cof),
        "0,00", "0,00", "0,00"])]
    for i, r in enumerate(grupo, 1):
        lines.append(to_pipe(["A170", str(i), gc(r,"COD_ITEM"), gc(r,"DESCR_COMPL"),
            fmt_valor_sped(r.get("VL_ITEM","")), fmt_valor_sped(r.get("VL_DESC_ITEM","")),
            gc(r,"NAT_BC_CRED"), gc(r,"IND_ORIG_CRED"), gc(r,"CST_PIS"),
            fmt_valor_sped(r.get("VL_BC_PIS_ITEM","")),
            fmt_valor_sped(r.get("ALIQ_PIS_ITEM","")),
            fmt_valor_sped(r.get("VL_PIS_ITEM","")),
            gc(r,"CST_COFINS"),
            fmt_valor_sped(r.get("VL_BC_COFINS_ITEM","")),
            fmt_valor_sped(r.get("ALIQ_COFINS_ITEM","")),
            fmt_valor_sped(r.get("VL_COFINS_ITEM","")),
            gc(r,"COD_CTA_ITEM"), gc(r,"COD_CCUS")]))
    return lines

def build_c_block(grupo: list) -> list:
    r0 = grupo[0]
    vl_doc  = soma_valores(*[r.get("VL_ITEM", 0) for r in grupo])
    vl_desc = soma_valores(*[r.get("VL_DESC_ITEM", 0) for r in grupo])
    vl_pis     = soma_valores(*[r.get("VL_PIS_ITEM", 0) for r in grupo])
    vl_cof     = soma_valores(*[r.get("VL_COFINS_ITEM", 0) for r in grupo])
    vl_icms    = soma_valores(*[r.get("VL_ICMS_ITEM", 0) for r in grupo])
    vl_ipi     = soma_valores(*[r.get("VL_IPI_ITEM", 0) for r in grupo])
    vl_bc_icms = soma_valores(*[r.get("VL_BC_ICMS_ITEM", 0) for r in grupo])
    # Layout C100 conforme Manual EFD Contribuicoes: 29 campos
    # REG|IND_OPER|IND_EMIT|COD_PART|COD_MOD|COD_SIT|SER|NUM_DOC|CHV_NFE|
    # DT_DOC|DT_E_S|VL_DOC|IND_PGTO|VL_DESC|VL_ABAT_NT|VL_MERC|IND_FRT|
    # VL_FRT|VL_SEG|VL_OUT_DA|VL_BC_ICMS|VL_ICMS|VL_BC_ICMS_ST|VL_ICMS_ST|
    # VL_IPI|VL_PIS|VL_COFINS|VL_PIS_ST|VL_COFINS_ST
    lines = [to_pipe(["C100",
        gc(r0,"IND_OPER"), gc(r0,"IND_EMIT"), gc(r0,"COD_PART\n(Fornecedor/Cliente)"),
        gc(r0,"COD_MOD"), gc(r0,"COD_SIT"), gc(r0,"SER"), gc(r0,"NUM_DOC"),
        gc(r0,"CHV_NFE"), fmt_data(r0.get("DT_DOC","")), fmt_data(r0.get("DT_E_S","")),
        fmt_br(vl_doc), gc(r0,"IND_PGTO"), fmt_br(vl_desc),
        fmt_br(soma_valores(r0.get("VL_ABAT_NT", 0))),
        fmt_br(vl_doc - vl_desc),
        gc(r0,"IND_FRT"),
        fmt_br(soma_valores(r0.get("VL_FRT", 0))),
        fmt_br(soma_valores(r0.get("VL_SEG", 0))),
        fmt_br(soma_valores(r0.get("VL_OUT_DA", 0))),
        fmt_br(vl_bc_icms), fmt_br(vl_icms),
        "0", "0",
        fmt_br(vl_ipi), fmt_br(vl_pis), fmt_br(vl_cof),
        "0", "0"])]
    for i, r in enumerate(grupo, 1):
        # Layout C170: 37 campos conforme Manual EFD Contribuicoes
        # IND_APUR: '0'=nao ha apuracao especifica (padrao), vazio invalido
        # QUANT_BC_PIS/COFINS e ALIQ_PIS_R$/COFINS_R$: '' quando nao se aplica
        # (usado apenas para tributacao por quantidade, CST 03)
        ind_apur = gc(r,"IND_APUR") or "0"
        # QUANT e ALIQ_R$: preenche so se CST usar quantidade (03)
        cst_pis    = gc(r,"CST_PIS")
        cst_cofins = gc(r,"CST_COFINS")
        quant_pis  = fmt_br(soma_valores(r.get("QUANT_BC_PIS",0))) if cst_pis == "03" else ""
        aliq_pis_r = fmt_br(soma_valores(r.get("ALIQ_PIS_R$",0))) if cst_pis == "03" else ""
        quant_cof  = fmt_br(soma_valores(r.get("QUANT_BC_COFINS",0))) if cst_cofins == "03" else ""
        aliq_cof_r = fmt_br(soma_valores(r.get("ALIQ_COFINS_R$",0))) if cst_cofins == "03" else ""
        lines.append(to_pipe(["C170", str(i), gc(r,"COD_ITEM"), gc(r,"DESCR_COMPL"),
            fmt_valor_sped(r.get("QTD","")), gc(r,"UNID"),
            fmt_valor_sped(r.get("VL_ITEM","")),
            fmt_valor_sped(r.get("VL_DESC_ITEM","")),
            gc(r,"IND_MOV"), gc(r,"CST_ICMS"),
            gc(r,"CFOP"), gc(r,"COD_NAT"),
            fmt_valor_sped(r.get("VL_BC_ICMS_ITEM","")),
            fmt_valor_sped(r.get("ALIQ_ICMS","")),
            fmt_valor_sped(r.get("VL_ICMS_ITEM","")),
            fmt_valor_sped(r.get("VL_BC_ICMS_ST_ITEM","")),
            fmt_valor_sped(r.get("ALIQ_ST","")),
            fmt_valor_sped(r.get("VL_ICMS_ST_ITEM","")),
            ind_apur,
            gc(r,"CST_IPI"), gc(r,"COD_ENQ","COD_NAT_IPI"),
            fmt_valor_sped(r.get("VL_BC_IPI","")),
            fmt_valor_sped(r.get("ALIQ_IPI","")),
            fmt_valor_sped(r.get("VL_IPI_ITEM","")),
            cst_pis,
            fmt_valor_sped(r.get("VL_BC_PIS_ITEM","")),
            fmt_valor_sped(r.get("ALIQ_PIS_%","")),
            quant_pis, aliq_pis_r,
            fmt_valor_sped(r.get("VL_PIS_ITEM","")),
            cst_cofins,
            fmt_valor_sped(r.get("VL_BC_COFINS_ITEM","")),
            fmt_valor_sped(r.get("ALIQ_COFINS_%","")),
            quant_cof, aliq_cof_r,
            fmt_valor_sped(r.get("VL_COFINS_ITEM","")),
            gc(r,"COD_CTA_ITEM")]))
    return lines


def build_c500_block(row) -> list:
    """
    Monta C500 + C501 + C505 a partir de uma linha da planilha C500_C501_C505.
    VL_PIS e VL_COFINS do C500 são preenchidos com os valores de C501/C505.
    Hierarquia: C500 (cabeçalho NF energia/comunicação/gás/água)
                C501 (PIS)
                C505 (COFINS)
    """
    vl_pis    = gc(row, "VL_PIS_501")
    vl_cofins = gc(row, "VL_COFINS_505")

    # Layout C500 conforme Manual EFD Contribuicoes v1.35:
    # REG|COD_PART|COD_MOD|COD_SIT|COD_SIT_ESP|IND_EMIT|NUM_DOC|
    # DT_DOC|DT_ENT|VL_DOC|VL_ICMS|COD_INF|VL_PIS|VL_COFINS|CHV_DOCe
    c500 = to_pipe(["C500",
        gc(row, "COD_PART"),
        gc(row, "COD_MOD"),
        gc(row, "COD_SIT"),
        gc(row, "COD_SIT_ESP"),
        gc(row, "IND_EMIT"),
        gc(row, "NUM_DOC"),
        fmt_data(row.get("DT_DOC", "")),
        fmt_data(row.get("DT_ENT", "")),
        fmt_valor_sped(row.get("VL_DOC", "")),
        fmt_valor_sped(row.get("VL_ICMS", "")),
        gc(row, "COD_INF"),
        fmt_valor_sped(vl_pis),
        fmt_valor_sped(vl_cofins),
        gc(row, "CHV_DOCe"),
    ])
    # Layout C501: REG|CST_PIS|VL_ITEM|NAT_BC_CRED|VL_BC_PIS|ALIQ_PIS|VL_PIS|COD_CTA
    c501 = to_pipe(["C501",
        gc(row, "CST_PIS"),
        fmt_valor_sped(row.get("VL_ITEM_C501", row.get("VL_BC_PIS", ""))),
        gc(row, "NAT_BC_CRED"),
        fmt_valor_sped(row.get("VL_BC_PIS", "")),
        fmt_valor_sped(row.get("ALIQ_PIS", "")),
        fmt_valor_sped(vl_pis),
        gc(row, "COD_CTA_501"),
    ])
    # Layout C505: REG|CST_COFINS|VL_ITEM|NAT_BC_CRED|VL_BC_COFINS|ALIQ_COFINS|VL_COFINS|COD_CTA
    c505 = to_pipe(["C505",
        gc(row, "CST_COFINS"),
        fmt_valor_sped(row.get("VL_ITEM_C505", row.get("VL_BC_COFINS", ""))),
        gc(row, "NAT_BC_CRED"),
        fmt_valor_sped(row.get("VL_BC_COFINS", "")),
        fmt_valor_sped(row.get("ALIQ_COFINS", "")),
        fmt_valor_sped(vl_cofins),
        gc(row, "COD_CTA_505"),
    ])
    return [c500, c501, c505]

def build_d_block(r) -> list:
    return [
        to_pipe(["D100", gc(r,"IND_OPER"), gc(r,"IND_EMIT"), gc(r,"COD_PART\n(Transportadora)"),
            gc(r,"COD_MOD"), gc(r,"COD_SIT"), gc(r,"SER"), gc(r,"SUB"), gc(r,"NUM_DOC"),
            gc(r,"CHV_CTE"), fmt_data(r.get("DT_DOC","")), fmt_data(r.get("DT_A_P","")), gc(r,"TP_CT_e"), gc(r,"CHV_CTE_REF"),
            fmt_valor(r.get("VL_DOC","")), fmt_valor(r.get("VL_DESC","")), gc(r,"IND_FRT"),
            fmt_valor(r.get("VL_SERV","")), fmt_valor(r.get("VL_BC_ICMS","")),
            fmt_valor(r.get("VL_ICMS","")), fmt_valor(r.get("VL_NT","")),
            gc(r,"COD_CTA"), gc(r,"COD_INF")]),
        to_pipe(["D101", gc(r,"IND_NAT_FRT"), fmt_valor(r.get("VL_ITEM_PIS","")),
            gc(r,"CST_PIS"), gc(r,"NAT_BC_CRED"), fmt_valor(r.get("VL_BC_PIS","")),
            fmt_valor(r.get("ALIQ_PIS_%","")), fmt_valor(r.get("QUANT_BC_PIS","")),
            fmt_valor(r.get("ALIQ_PIS_R$","")), fmt_valor(r.get("VL_PIS","")),
            gc(r,"COD_CTA_PIS"), gc(r,"COD_CCUS_PIS")]),
        to_pipe(["D105", gc(r,"IND_NAT_FRT"), fmt_valor(r.get("VL_ITEM_COF","")),
            gc(r,"CST_COFINS"), gc(r,"NAT_BC_CRED"), fmt_valor(r.get("VL_BC_COFINS","")),
            fmt_valor(r.get("ALIQ_COFINS_%","")), fmt_valor(r.get("QUANT_BC_COF","")),
            fmt_valor(r.get("ALIQ_COFINS_R$","")), fmt_valor(r.get("VL_COFINS","")),
            gc(r,"COD_CTA_COF"), gc(r,"COD_CCUS_COF")]),
    ]

def build_f100(r):
    # Layout F100 conforme Manual EFD Contribuicoes v1.35 (19 campos):
    # REG|IND_OPER|COD_PART|COD_ITEM|DT_OPER|VL_OPER|CST_PIS|VL_BC_PIS|ALIQ_PIS|VL_PIS|
    # CST_COFINS|VL_BC_COFINS|ALIQ_COFINS|VL_COFINS|NAT_BC_CRED|IND_ORIG_CRED|
    # COD_CTA|COD_CCUS|DESC_DOC_OPER
    return to_pipe(["F100",
        gc(r,"IND_OPER"),
        gc(r,"COD_PART"),
        gc(r,"COD_ITEM"),
        fmt_data(r.get("DT_OPER","")),
        fmt_valor_sped(r.get("VL_OPER","")),
        gc(r,"CST_PIS"),
        fmt_valor_sped(r.get("VL_BC_PIS","")),
        fmt_valor_sped(r.get("ALIQ_PIS_%", r.get("ALIQ_PIS",""))),
        fmt_valor_sped(r.get("VL_PIS","")),
        gc(r,"CST_COFINS"),
        fmt_valor_sped(r.get("VL_BC_COFINS","")),
        fmt_valor_sped(r.get("ALIQ_COFINS_%", r.get("ALIQ_COFINS",""))),
        fmt_valor_sped(r.get("VL_COFINS","")),
        gc(r,"NAT_BC_CRED"),
        gc(r,"IND_ORIG_CRED"),
        gc(r,"COD_CTA"),
        gc(r,"COD_CCUS"),
        gc(r,"DES_COMPL"),
    ])

def build_f120(r):
    # Layout F120 conforme Manual EFD Contribuicoes v1.35 (18 campos):
    # REG|NAT_BC_CRED|IDENT_BEM_IMOB|IND_ORIG_CRED|IND_UTIL_BEM_IMOB|
    # VL_OPER_DEP|PARC_OPER_NAO_BC_CRED|CST_PIS|VL_BC_PIS|ALIQ_PIS|VL_PIS|
    # CST_COFINS|VL_BC_COFINS|ALIQ_COFINS|VL_COFINS|COD_CTA|COD_CCUS|DESC_BEM_IMOB
    return to_pipe(["F120",
        gc(r,"NAT_BC_CRED"),
        gc(r,"IDENT_BEM_IMOB"),
        gc(r,"IND_ORIG_CRED"),
        gc(r,"IND_UTIL_BEM_IMOB"),
        fmt_valor_sped(r.get("VL_OPER_DEP","")),
        fmt_valor_sped(r.get("PARC_NAO_BC", r.get("PARC_OPER_NAO_BC_CRED",""))),
        gc(r,"CST_PIS"),
        fmt_valor_sped(r.get("VL_BC_PIS","")),
        fmt_valor_sped(r.get("ALIQ_PIS_%", r.get("ALIQ_PIS",""))),
        fmt_valor_sped(r.get("VL_PIS","")),
        gc(r,"CST_COFINS"),
        fmt_valor_sped(r.get("VL_BC_COFINS","")),
        fmt_valor_sped(r.get("ALIQ_COFINS_%", r.get("ALIQ_COFINS",""))),
        fmt_valor_sped(r.get("VL_COFINS","")),
        gc(r,"COD_CTA"),
        gc(r,"COD_CCUS"),
        gc(r,"DES_COMPL"),
    ])

def build_f130(r):
    # Layout F130 conforme Manual EFD Contribuicoes v1.35 (21 campos):
    # REG|NAT_BC_CRED|IDENT_BEM_IMOB|IND_ORIG_CRED|IND_UTIL_BEM_IMOB|
    # MES_OPER_AQUIS|VL_OPER_AQUIS|PARC_OPER_NAO_BC_CRED|VL_BC_CRED|IND_NR_PARC|
    # CST_PIS|VL_BC_PIS|ALIQ_PIS|VL_PIS|CST_COFINS|VL_BC_COFINS|ALIQ_COFINS|VL_COFINS|
    # COD_CTA|COD_CCUS|DESC_BEM_IMOB
    return to_pipe(["F130",
        gc(r,"NAT_BC_CRED"),
        gc(r,"IDENT_BEM_IMOB"),
        gc(r,"IND_ORIG_CRED"),
        gc(r,"IND_UTIL_BEM_IMOB"),
        gc(r,"MES_OPER_AQUIS"),
        fmt_valor_sped(r.get("VL_OPER_AQUIS","")),
        fmt_valor_sped(r.get("PARC_NAO_BC", r.get("PARC_OPER_NAO_BC_CRED",""))),
        fmt_valor_sped(r.get("VL_BC_CRED","")),
        gc(r,"IND_NR_PARC"),
        gc(r,"CST_PIS"),
        fmt_valor_sped(r.get("VL_BC_PIS","")),
        fmt_valor_sped(r.get("ALIQ_PIS_%", r.get("ALIQ_PIS",""))),
        fmt_valor_sped(r.get("VL_PIS","")),
        gc(r,"CST_COFINS"),
        fmt_valor_sped(r.get("VL_BC_COFINS","")),
        fmt_valor_sped(r.get("ALIQ_COFINS_%", r.get("ALIQ_COFINS",""))),
        fmt_valor_sped(r.get("VL_COFINS","")),
        gc(r,"COD_CTA"),
        gc(r,"COD_CCUS"),
        gc(r,"DES_COMPL"),
    ])


# ─────────────────────────────────────────────────────────────────────────────
# Bloco 9 e injeção
# ─────────────────────────────────────────────────────────────────────────────
def recalc_totalizadores(lines: list) -> list:
    """
    Recalcula os registros X990 de cada bloco (0990, A990, C990, D990, F990, etc).
    
    Regra do Manual EFD Contribuicoes:
    - X990 contem o total de linhas do bloco incluindo o proprio X990
    - Cada bloco eh delimitado pelo X001 (abertura) e X990 (encerramento)
    - O bloco 9 e calculado separadamente pelo recalc_9900
    
    Algoritmo: para cada bloco, conta as linhas entre X001 e X990 inclusive.
    """
    # Mapeia posicao de cada X001 e X990
    # Estrutura: {bloco: (idx_001, idx_990)}
    blocos = {}
    for i, line in enumerate(lines):
        ls = line.strip()
        if not ls.startswith("|"):
            continue
        reg = parse_pipe(ls)[0]
        if not reg:
            continue
        # X001 = abertura do bloco (ex: A001, C001, D001, F001, 0001)
        if len(reg) == 4 and reg.endswith("001"):
            bloco = reg[0]
            if bloco not in blocos:
                blocos[bloco] = {"ini": i, "fim": None}
        # 0000 = primeiro registro do bloco 0 (sem X001 proprio)
        if reg == "0000" and "0" not in blocos:
            blocos["0"] = {"ini": i, "fim": None}
        # X990 = encerramento do bloco
        if len(reg) == 4 and reg.endswith("990"):
            bloco = reg[0]
            if bloco in blocos:
                blocos[bloco]["fim"] = i
            else:
                blocos[bloco] = {"ini": None, "fim": i}

    # Calcula contagem real de cada bloco (ini ate fim inclusive)
    contadores = {}
    for bloco, pos in blocos.items():
        if pos["ini"] is not None and pos["fim"] is not None:
            # Conta todas as linhas de registro entre abertura e fechamento
            count = 0
            for i in range(pos["ini"], pos["fim"] + 1):
                if lines[i].strip().startswith("|"):
                    count += 1
            contadores[bloco] = count
        elif pos["fim"] is not None:
            # Sem X001 (bloco 0): conta do inicio ate o X990
            count = 0
            for i in range(0, pos["fim"] + 1):
                if lines[i].strip().startswith("|"):
                    count += 1
            contadores[bloco] = count

    # Atualiza as linhas X990 com o valor correto
    result = []
    for line in lines:
        ls = line.strip()
        if ls.startswith("|"):
            f = parse_pipe(ls)
            reg = f[0] if f else ""
            # Nao recalcula 9990 aqui (feito pelo recalc_9900)
            if len(reg) == 4 and reg.endswith("990") and reg != "9990":
                bloco = reg[0]
                if bloco in contadores:
                    result.append(to_pipe([reg, str(contadores[bloco])]))
                    continue
        result.append(line)
    return result

def recalc_9900(lines: list) -> list:
    counter = defaultdict(int)
    new_lines = []
    skip = False
    for line in lines:
        ls = line.strip()
        if re.match(r"^\|(9001|9900|9990|9999)\|", ls):
            skip = True
            continue
        if skip:
            continue
        new_lines.append(line)
        if ls.startswith("|"):
            reg = parse_pipe(ls)[0]
            counter[reg] += 1
    counter["9001"] = 1
    # Monta registros 9900 (um por tipo de registro)
    linhas_9900 = []
    for reg in sorted(counter.keys()):
        linhas_9900.append(f"|9900|{reg}|{counter[reg]}|\n")
    # Quantidade de linhas 9900 inclui o proprio |9900|9900|...|
    n_9900 = len(linhas_9900) + 1
    linhas_9900.append(f"|9900|9900|{n_9900}|\n")
    linhas_9900.append("|9900|9990|1|\n")
    linhas_9900.append("|9900|9999|1|\n")

    # Bloco 9: 9001 + todos os 9900 + 9990 + 9999
    # 9990 conta todas as linhas do bloco 9 incluindo ele mesmo
    n_bloco9 = 1 + len(linhas_9900) + 1 + 1  # 9001 + 9900s + 9990 + 9999
    bloco9 = ["|9001|0|\n"]
    bloco9.extend(linhas_9900)
    bloco9.append(f"|9990|{n_bloco9}|\n")

    # 9999: total geral de linhas do arquivo incluindo o proprio 9999
    total_arquivo = len(new_lines) + len(bloco9) + 1
    bloco9.append(f"|9999|{total_arquivo}|\n")
    return new_lines + bloco9

def find_x010_ranges(lines: list, prefix: str) -> dict:
    ranges = {}
    positions = []
    x990 = None
    for i, line in enumerate(lines):
        ls = line.strip()
        if ls.startswith(f"|{prefix}010|"):
            f = parse_pipe(ls)
            positions.append((i, f[1] if len(f) > 1 else ""))
        if ls.startswith(f"|{prefix}990|"):
            x990 = i
    for k, (pos, cnpj) in enumerate(positions):
        fim = positions[k+1][0] if k+1 < len(positions) else (x990 or len(lines))
        ranges[cnpj] = fim
    return ranges

def inject_by_cnpj(lines: list, prefix: str, cnpj_map: dict, log_lines: list) -> list:
    """
    Injeta registros dentro do sub-bloco X010 de cada estabelecimento.

    Regra fundamental: o CNPJ_ESTAB da planilha determina em qual
    sub-bloco X010 o registro sera inserido. Nunca mistura registros
    de estabelecimentos diferentes.

    Algoritmo: processa as insercoes em ordem DECRESCENTE de posicao.
    Isso garante que cada insercao nao desloca as posicoes das
    insercoes subsequentes (que estao em posicoes menores).
    """
    ranges = find_x010_ranges(lines, prefix)

    # Monta lista de (posicao, cnpj, linhas) e ordena por posicao DECRESCENTE
    insercoes = []
    for cnpj, novas in cnpj_map.items():
        if not novas:
            continue
        if cnpj not in ranges:
            log_lines.append(f"⚠ CNPJ {cnpj} sem {prefix}010 no TXT — pulando")
            continue
        insercoes.append((ranges[cnpj], cnpj, novas))

    # Ordena por posicao decrescente: insere do fim para o inicio
    # Assim cada insercao nao afeta as posicoes das anteriores
    insercoes.sort(key=lambda x: x[0], reverse=True)

    for pos, cnpj, novas in insercoes:
        lines[pos:pos] = novas
        log_lines.append(f"✔ Bloco {prefix}: {len(novas)} linha(s) → CNPJ {cnpj}")

    return lines

def agrupar_por_doc(rows: list) -> dict:
    """
    Agrupa linhas de itens por documento fiscal.

    Chave de agrupamento:
      CNPJ_ESTAB + PERIODO + NUM_DOC + COD_PART + CHV_DOC

    O CNPJ_ESTAB e a chave principal — cada documento e inserido
    SOMENTE no sub-bloco do estabelecimento correto (A010/C010/D010).

    COD_PART e CHV_DOC entram na chave para evitar mistura entre
    documentos de fornecedores diferentes que tenham o mesmo numero
    de NF (situacao comum em multi-estabelecimentos).
    """
    grupos = {}
    for row in rows:
        cnpj     = gc(row, "CNPJ_ESTAB")
        periodo  = str(gc(row, "PERÍODO\n(MMAAAA)", "PERÍODO", "PERIODO")).strip()
        ndoc     = str(gc(row, "NUM_DOC")).strip()
        cod_part = str(gc(row,
                         "COD_PART\n(Fornecedor/Cliente)",
                         "COD_PART\n(Transportadora)",
                         "COD_PART")).strip()
        chv      = str(gc(row, "CHV_DOC", "CHV_NFE", "CHV_NFSE", "CHV_CTE")).strip()

        if not cnpj or not ndoc:
            continue

        # Chave: CNPJ + periodo + numero doc + fornecedor + chave eletronica
        key = f"{cnpj}||{periodo}||{ndoc}||{cod_part}||{chv}"

        if key not in grupos:
            grupos[key] = {"cnpj": cnpj, "rows": []}
        grupos[key]["rows"].append(row)
    return grupos


# ─────────────────────────────────────────────────────────────────────────────
# Motor principal
# ─────────────────────────────────────────────────────────────────────────────
def retificar(txt_bytes: bytes, planilha: dict) -> tuple[bytes, list]:
    log = []
    enc = detect_encoding(txt_bytes)
    lines = txt_bytes.decode(enc, errors="replace").splitlines(keepends=True)
    periodo = get_periodo(lines)
    log.append(f"→ Período detectado: {periodo or 'não encontrado'}")
    parsed = parse_txt(lines)

    # 0000 → IND_RET = 1
    for i, line in enumerate(lines):
        if line.strip().startswith("|0000|"):
            f = parse_pipe(line.strip())
            f[2] = "1"
            lines[i] = to_pipe(f)
            log.append("✔ 0000 marcado como retificador (IND_RET=1)")
            break

    # ── Hierarquia do bloco 0 conforme Manual EFD Contribuições ─────────────
    # Ordem: 0140 → 0150 → 0190 → 0200 → 0400 → 0450 → 0500 → 0990
    # Cada novo registro é inserido APÓS os registros existentes do mesmo tipo,
    # mantendo a hierarquia e nunca ultrapassando o 0990.

    def find_last_reg(lines_list, reg_prefix, before_idx):
        """Encontra o índice da última linha do registro, limitado a before_idx."""
        last = None
        for i, line in enumerate(lines_list):
            if i >= before_idx:
                break
            if line.strip().startswith(f"|{reg_prefix}|"):
                last = i
        return last

    b0end = parsed["bloco_0_end"]
    offset = 0  # acumula deslocamento a cada inserção

    def insert_after_last(reg_prefix, new_lines):
        nonlocal offset
        if not new_lines:
            return
        b0 = b0end + offset
        last_pos = find_last_reg(lines, reg_prefix, b0)
        if last_pos is not None:
            insert_at = last_pos + 1
        else:
            # Não existe nenhum registro desse tipo: insere antes do 0990
            insert_at = b0
        lines[insert_at:insert_at] = new_lines
        offset += len(new_lines)

    # 0140 — novos estabelecimentos
    novos_0140 = []
    for emp in filter_periodo(planilha.get("0_EMPRESA", []), periodo):
        cnpj = gc(emp, "CNPJ_ESTAB", "CNPJ")
        if not cnpj:
            continue
        if cnpj in parsed["cnpj_0140"]:
            log.append(f"→ 0140 CNPJ {cnpj} já existe — mantém")
        else:
            novos_0140.append(build_0140(emp))
            log.append(f"✔ 0140 CNPJ {cnpj} criado")
    insert_after_last("0140", novos_0140)

    # 0150 — participantes (filtrado por período + sem duplicar COD_PART)
    cod_0150_existentes = set()
    for line in lines:
        if line.strip().startswith("|0150|"):
            f = parse_pipe(line.strip())
            cod_0150_existentes.add(f[1] if len(f) > 1 else "")
    novos_0150 = []
    vistos_0150 = set()
    for r in filter_periodo(planilha.get("0150_PARTICIPANTES", []), periodo):
        cod = gc(r, "COD_PART")
        if cod and cod not in cod_0150_existentes and cod not in vistos_0150:
            novos_0150.append(build_0150(r))
            vistos_0150.add(cod)
    insert_after_last("0150", novos_0150)

    # 0190 — unidades (filtrado por período + sem duplicar UNID)
    ex_0190 = set()
    for line in lines:
        if line.strip().startswith("|0190|"):
            f = parse_pipe(line.strip()); ex_0190.add(f[1] if len(f) > 1 else "")
    novos_0190 = []
    vistos_0190 = set()
    for r in filter_periodo(planilha.get("0190_UNIDADES", []), periodo):
        u = gc(r, "UNID")
        if u and u not in ex_0190 and u not in vistos_0190:
            novos_0190.append(build_0190(r))
            vistos_0190.add(u)
    insert_after_last("0190", novos_0190)

    # 0200 — itens (filtrado por período + sem duplicar COD_ITEM)
    novos_0200 = []
    for r in filter_periodo(planilha.get("0200_ITENS", []), periodo):
        cod = gc(r, "COD_ITEM")
        if not cod:
            continue
        if cod not in parsed["cod_0200"]:
            novos_0200.append(build_0200(r))
            parsed["cod_0200"].add(cod)
            log.append(f"✔ 0200 item {cod} inserido")
        else:
            log.append(f"→ 0200 item {cod} já existe — mantém")
    insert_after_last("0200", novos_0200)

    # 0400 — naturezas de receita (filtrado por período + sem duplicar COD_NAT_REC)
    ex_0400 = set()
    for line in lines:
        if line.strip().startswith("|0400|"):
            f = parse_pipe(line.strip()); ex_0400.add(f[1] if len(f) > 1 else "")
    novos_0400 = []
    vistos_0400 = set()
    for r in filter_periodo(planilha.get("0400_NAT_REC", []), periodo):
        c = gc(r, "COD_NAT_REC")
        if c and c not in ex_0400 and c not in vistos_0400:
            novos_0400.append(build_0400(r))
            vistos_0400.add(c)
    insert_after_last("0400", novos_0400)

    # 0500 — plano de contas (filtrado por período + sem duplicar COD_CTA)
    ex_0500 = set()
    for line in lines:
        if line.strip().startswith("|0500|"):
            f = parse_pipe(line.strip())
            ex_0500.add(f[5] if len(f) > 5 else "")  # COD_CTA está na posição 5
    novos_0500 = []
    vistos_0500 = set()
    for r in filter_periodo(planilha.get("0500_PLANO_CONTAS", []), periodo):
        c = gc(r, "COD_CTA")
        if c and c not in ex_0500 and c not in vistos_0500:
            novos_0500.append(build_0500(r))
            vistos_0500.add(c)
    insert_after_last("0500", novos_0500)

    total_ins = len(novos_0140) + len(novos_0150) + len(novos_0190) + len(novos_0200) + len(novos_0400) + len(novos_0500)
    if total_ins:
        log.append(f"✔ {total_ins} linha(s) inseridas no bloco 0 (ordem hierárquica)") 

    # Bloco A
    a_map = defaultdict(list)
    for g in agrupar_por_doc(filter_periodo(planilha.get("A100_A170", []), periodo)).values():
        a_map[g["cnpj"]].extend(build_a_block(g["rows"]))
    lines = inject_by_cnpj(lines, "A", dict(a_map), log)

    # Bloco C — C100/C170 (NF-e) e C500/C501/C505 (energia/comunicação/gás/água)
    # Ambos injetados dentro do C010 do estabelecimento correspondente
    c_map = defaultdict(list)
    for g in agrupar_por_doc(filter_periodo(planilha.get("C100_C170", []), periodo)).values():
        c_map[g["cnpj"]].extend(build_c_block(g["rows"]))
    for r in filter_periodo(planilha.get("C500_C501_C505", []), periodo):
        cnpj = gc(r, "CNPJ_ESTAB")
        if cnpj:
            c_map[cnpj].extend(build_c500_block(r))
    lines = inject_by_cnpj(lines, "C", dict(c_map), log)

    # Bloco D
    d_map = defaultdict(list)
    for r in filter_periodo(planilha.get("D100_D101_D105", []), periodo):
        cnpj = gc(r, "CNPJ_ESTAB")
        if cnpj:
            d_map[cnpj].extend(build_d_block(r))
    lines = inject_by_cnpj(lines, "D", dict(d_map), log)

    # Bloco F — F100, F120 e F130 dentro do F010 do estabelecimento correto
    # Ordem dentro do sub-bloco F010: F100 → F120 → F130
    f_map = defaultdict(list)
    for aba, fn in [("F100", build_f100), ("F120", build_f120), ("F130", build_f130)]:
        for r in filter_periodo(planilha.get(aba, []), periodo):
            cnpj = gc(r, "CNPJ_ESTAB")
            if cnpj:
                f_map[cnpj].append(fn(r))
    lines = inject_by_cnpj(lines, "F", dict(f_map), log)

    lines = recalc_totalizadores(lines)
    lines = recalc_9900(lines)
    log.append(f"✔ Totalizadores e bloco 9 recalculados — {len(lines)} linhas no total")

    result_bytes = "".join(lines).encode(enc, errors="replace")
    return result_bytes, log


# ─────────────────────────────────────────────────────────────────────────────
# Interface Streamlit
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("## 📄 EFD Contribuições — Retificação")
st.markdown("Faça upload da planilha modelo e dos TXTs originais. O sistema processa cada período automaticamente e gera os arquivos retificados para download.")
st.divider()

col1, col2 = st.columns(2)

with col1:
    st.markdown("### 1. Planilha modelo")
    xlsx_file = st.file_uploader(
        "Arquivo .xlsx com os novos registros",
        type=["xlsx"],
        help="Use o modelo EFD_Contribuicoes_Modelo_v3.xlsx"
    )

with col2:
    st.markdown("### 2. TXTs originais")
    txt_files = st.file_uploader(
        "Um ou mais arquivos .txt (um por período)",
        type=["txt"],
        accept_multiple_files=True,
        help="Ex.: EFD_012024.txt, EFD_022024.txt"
    )

st.divider()

# Regras
with st.expander("📋 Regras aplicadas automaticamente", expanded=False):
    r1, r2, r3 = st.columns(3)
    with r1:
        st.markdown("**0140 — Estabelecimentos**")
        st.caption("CNPJ já existente no TXT é reutilizado. Novo 0140 criado apenas se não encontrado.")
        st.markdown("**0200 — Itens centralizados**")
        st.caption("Novo item inserido uma vez no bloco 0. COD_ITEM duplicado é ignorado.")
    with r2:
        st.markdown("**A100 / C100 — Totais automáticos**")
        st.caption("VL_DOC, VL_PIS, VL_COFINS calculados somando os itens A170 / C170.")
        st.markdown("**Bloco 9 — Contadores**")
        st.caption("Todos os registros 9900 recalculados. 0000 marcado IND_RET = 1.")
    with r3:
        st.markdown("**0150 / 0190 / 0400 / 0500**")
        st.caption("Inseridos sem duplicar registros já presentes no TXT original.")
        st.markdown("**Blocos A / C / D / F**")
        st.caption("Cada documento inserido no sub-bloco A010 / C010 do CNPJ correto.")

# Botão processar
if xlsx_file and txt_files:
    m1, m2, m3 = st.columns(3)
    m1.metric("Planilha", xlsx_file.name)
    m2.metric("TXTs enviados", len(txt_files))
    m3.metric("Arquivos a gerar", len(txt_files))
    st.divider()

    if st.button("🚀 Processar retificação", type="primary", use_container_width=True):
        with st.spinner("Lendo planilha…"):
            try:
                planilha = load_planilha(xlsx_file.read())
                st.success(f"Planilha lida — {len(planilha)} abas carregadas")
            except Exception as e:
                st.error(f"Erro ao ler planilha: {e}")
                st.stop()

        resultados = []
        todos_logs = []

        progress = st.progress(0, text="Iniciando…")

        for i, txt_file in enumerate(txt_files):
            nome_saida = txt_file.name.replace(".txt", "_RET.txt").replace(".TXT", "_RET.txt")
            st.markdown(f"**Processando: {txt_file.name}**")

            try:
                result_bytes, log_lines = retificar(txt_file.read(), planilha)
                resultados.append((nome_saida, result_bytes))
                todos_logs.extend([f"[{txt_file.name}] {l}" for l in log_lines])

                log_html = "<br>".join(
                    f'<span class="{"status-ok" if l.startswith("✔") else "status-warn" if l.startswith("⚠") else "status-info"}">{l}</span>'
                    for l in log_lines
                )
                st.markdown(f'<div class="log-box">{log_html}</div>', unsafe_allow_html=True)
                st.success(f"✔ {nome_saida} gerado — {len(result_bytes):,} bytes")

            except Exception as e:
                st.error(f"Erro ao processar {txt_file.name}: {e}")
                todos_logs.append(f"[{txt_file.name}] ✖ Erro: {e}")

            progress.progress((i + 1) / len(txt_files), text=f"{i+1}/{len(txt_files)} arquivo(s) processado(s)")

        progress.progress(1.0, text="Concluído!")
        st.divider()
        st.markdown("### 📥 Downloads")

        if len(resultados) == 1:
            nome, dados = resultados[0]
            st.download_button(
                label=f"⬇ Baixar {nome}",
                data=dados,
                file_name=nome,
                mime="text/plain",
                use_container_width=True,
            )
        elif len(resultados) > 1:
            # ZIP com todos
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                for nome, dados in resultados:
                    zf.writestr(nome, dados)
            zip_buffer.seek(0)
            st.download_button(
                label=f"⬇ Baixar todos ({len(resultados)} arquivos em .zip)",
                data=zip_buffer,
                file_name="EFD_Retificados.zip",
                mime="application/zip",
                use_container_width=True,
            )
            st.markdown("Ou baixe individualmente:")
            for nome, dados in resultados:
                st.download_button(
                    label=f"⬇ {nome}",
                    data=dados,
                    file_name=nome,
                    mime="text/plain",
                    key=nome,
                )
else:
    st.info("Faça upload da planilha .xlsx e de um ou mais TXTs originais para habilitar o processamento.")

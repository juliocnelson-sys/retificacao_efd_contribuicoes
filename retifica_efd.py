"""
EFD Contribuições – Script de Retificação
==========================================
Lê a planilha modelo + TXTs originais e injeta os novos registros,
respeitando as seguintes regras:

Regras de 0140:
  - Se o CNPJ_ESTAB da planilha já existe em algum 0140 do TXT → reutiliza, sem duplicar.
  - Se não existe → cria novo 0140 com os dados da aba 0_EMPRESA.

Regras de 0200:
  - Centralizado: os itens ficam no mesmo 0140 existente (não cria duplicata).
  - Só insere item se COD_ITEM ainda não existir no bloco 0200 do TXT.

Regras de totais (A100/C100):
  - O script agrupa as linhas por CNPJ_ESTAB + PERÍODO + NUM_DOC.
  - Calcula VL_DOC, VL_PIS, VL_COFINS, VL_BC_PIS, VL_BC_COFINS somando os itens filhos.

Ao final:
  - Recalcula todos os contadores 9900.
  - Marca IND_RET = 1 no registro 0000.
  - Salva novo TXT com sufixo _RET.
"""

import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
ENCODING = "latin-1"
SEP = "|"

def parse_pipe(line: str) -> list:
    """Divide uma linha SPED em campos (remove pipes externos)."""
    return line.strip().strip("|").split("|")

def to_pipe(fields: list) -> str:
    """Monta linha SPED a partir de lista de campos."""
    return "|" + "|".join(str(f) for f in fields) + "|\n"

def fmt_valor(v) -> str:
    """Normaliza valor numérico para string brasileira."""
    if v is None or str(v).strip() == "":
        return "0,00"
    s = str(v).strip().replace(".", ",")
    if "," not in s:
        s += ",00"
    return s

def soma_valores(*vals) -> float:
    """Soma valores no formato brasileiro."""
    total = 0.0
    for v in vals:
        s = str(v).strip().replace(".", "").replace(",", ".")
        try:
            total += float(s)
        except ValueError:
            pass
    return total

def fmt_br(f: float) -> str:
    return f"{f:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def periodo_to_dates(periodo: str):
    """'102021' → ('01102021','31102021') – simplificado por mês."""
    import calendar
    mm = int(periodo[:2])
    aaaa = int(periodo[2:])
    last = calendar.monthrange(aaaa, mm)[1]
    return f"01{periodo}", f"{last:02d}{periodo}"

def read_txt(path: str) -> list:
    return Path(path).read_text(encoding=ENCODING).splitlines(keepends=True)

def write_txt(path: str, lines: list):
    Path(path).write_text("".join(lines), encoding=ENCODING)

# ─────────────────────────────────────────────────────────────────────────────
# Leitura da planilha
# ─────────────────────────────────────────────────────────────────────────────
def sheet_to_dicts(ws) -> list[dict]:
    """Lê uma aba do Excel e retorna lista de dicts {header: valor}."""
    headers = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # linha de cabeçalho real: primeira linha não-vazia que tem conteúdo na col 1
        if i == 0:
            continue  # título
        if not headers:
            headers = [str(c).strip() if c else f"COL{j}" for j, c in enumerate(row)]
            continue
        if all(c is None for c in row):
            continue
        rows.append({headers[j]: (row[j] if row[j] is not None else "")
                     for j in range(len(headers))})
    return rows

def load_planilha(xlsx_path: str) -> dict:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    data = {}
    for name in wb.sheetnames:
        if name in ("INSTRUÇÕES", "LISTAS"):
            continue
        data[name] = sheet_to_dicts(wb[name])
    return data

# ─────────────────────────────────────────────────────────────────────────────
# Parse do TXT original
# ─────────────────────────────────────────────────────────────────────────────
def parse_txt(lines: list) -> dict:
    """
    Retorna estrutura com:
      meta        – campos do 0000
      cnpj_0140   – {cnpj: fields_0140}   (para lookup rápido)
      cod_0140    – {cod_estab: cnpj}
      cod_0200    – set de COD_ITEM já existentes
      bloco_0_end – índice da linha 0990
      blocos      – {bloco: [(idx_inicio, idx_fim)]} para A,C,D,F
      raw         – as linhas originais
    """
    result = {
        "meta": {},
        "cnpj_0140": {},   # cnpj → fields list
        "cod_0140": {},    # cod_estab → cnpj
        "cod_0200": set(),
        "bloco_0_end": None,
        "raw": lines,
    }
    for idx, line in enumerate(lines):
        line_s = line.strip()
        if not line_s.startswith("|"):
            continue
        f = parse_pipe(line_s)
        reg = f[0]

        if reg == "0000":
            result["meta"] = f

        elif reg == "0140":
            # |0140|COD_ESTAB|NOME|CNPJ|UF|IE|COD_MUN|IM|SUFRAMA|
            cod  = f[1] if len(f) > 1 else ""
            cnpj = f[3] if len(f) > 3 else ""
            result["cnpj_0140"][cnpj] = f
            result["cod_0140"][cod]   = cnpj

        elif reg == "0200":
            cod_item = f[1] if len(f) > 1 else ""
            result["cod_0200"].add(cod_item)

        elif reg == "0990":
            result["bloco_0_end"] = idx

    return result

# ─────────────────────────────────────────────────────────────────────────────
# Montagem de novos registros
# ─────────────────────────────────────────────────────────────────────────────
def build_0140(row: dict) -> str:
    """Monta linha 0140 a partir de linha da aba 0_EMPRESA."""
    cnpj  = str(row.get("CNPJ", "")).strip()
    nome  = str(row.get("NOME / RAZÃO SOCIAL", "")).strip()
    uf    = str(row.get("UF", "")).strip()
    mun   = str(row.get("COD_MUN", "")).strip()
    # usa CNPJ como COD_ESTAB quando não há código específico
    cod   = cnpj
    return to_pipe(["0140", cod, nome, cnpj, uf, "", mun, "", ""])

def build_0200(row: dict) -> str:
    f = ["0200",
         str(row.get("COD_ITEM","")).strip(),
         str(row.get("DESCR_ITEM","")).strip(),
         str(row.get("COD_BARRA","")).strip(),
         "",   # COD_ANT_ITEM
         str(row.get("UNID_INV","")).strip(),
         str(row.get("TIPO_ITEM","")).strip(),
         str(row.get("COD_NCM","")).strip(),
         str(row.get("EX_IPI","")).strip(),
         str(row.get("COD_GEN","")).strip(),
         str(row.get("COD_LST","")).strip(),
         fmt_valor(row.get("ALIQ_ICMS","")),
    ]
    return to_pipe(f)

def build_0150(row: dict) -> str:
    f = ["0150",
         str(row.get("COD_PART","")).strip(),
         str(row.get("NOME","")).strip(),
         str(row.get("COD_PAIS","1058")).strip(),
         str(row.get("CNPJ","")).strip(),
         str(row.get("CPF","")).strip(),
         str(row.get("IE","")).strip(),
         str(row.get("COD_MUN","")).strip(),
         str(row.get("SUFRAMA","")).strip(),
         str(row.get("END","")).strip(),
         str(row.get("NUM","")).strip(),
         str(row.get("COMPL","")).strip(),
         str(row.get("BAIRRO","")).strip(),
    ]
    return to_pipe(f)

def build_0190(row: dict) -> str:
    return to_pipe(["0190",
                    str(row.get("UNID","")).strip(),
                    str(row.get("DESCR","")).strip()])

def build_0400(row: dict) -> str:
    return to_pipe(["0400",
                    str(row.get("COD_NAT_REC","")).strip(),
                    str(row.get("DESCR_NAT_REC","")).strip()])

def build_0500(row: dict) -> str:
    return to_pipe(["0500",
                    str(row.get("DT_ALT","")).strip(),
                    str(row.get("COD_CTA","")).strip(),
                    str(row.get("NÍVEL","")).strip(),
                    str(row.get("IND_CTA","")).strip(),
                    str(row.get("NOME_CTA","")).strip(),
                    str(row.get("COD_CTA_SUP","")).strip(),
                    str(row.get("COD_CTA_REF","")).strip(),
                    ""])

# ── A100/A170 ────────────────────────────────────────────────────────────────
def build_a_block(grupo: list) -> list:
    """
    grupo: lista de linhas da planilha A100_A170 que pertencem ao mesmo documento.
    Retorna lista de strings (A100 + N×A170).
    """
    r0 = grupo[0]
    # totais calculados somando os itens
    vl_doc     = soma_valores(*[r.get("VL_ITEM","0") for r in grupo])
    vl_desc    = soma_valores(*[r.get("VL_DESC_ITEM","0") for r in grupo])
    vl_bc_pis  = soma_valores(*[r.get("VL_BC_PIS_ITEM","0") for r in grupo])
    vl_pis     = soma_valores(*[r.get("VL_PIS_ITEM","0") for r in grupo])
    vl_bc_cof  = soma_valores(*[r.get("VL_BC_COFINS_ITEM","0") for r in grupo])
    vl_cof     = soma_valores(*[r.get("VL_COFINS_ITEM","0") for r in grupo])
    # alíquota: pega do primeiro item
    aliq_pis   = fmt_valor(r0.get("ALIQ_PIS_ITEM",""))
    aliq_cof   = fmt_valor(r0.get("ALIQ_COFINS_ITEM",""))

    a100 = to_pipe(["A100",
        str(r0.get("IND_OPER","")).strip(),
        str(r0.get("IND_EMIT","")).strip(),
        str(r0.get("COD_PART\n(Fornecedor/Cliente)","")).strip(),
        str(r0.get("COD_MOD","")).strip(),
        str(r0.get("COD_SIT","")).strip(),
        str(r0.get("SER","")).strip(),
        str(r0.get("SUB","")).strip(),
        str(r0.get("NUM_DOC","")).strip(),
        str(r0.get("CHV_DOC","")).strip(),
        str(r0.get("DT_DOC","")).strip(),
        str(r0.get("DT_EXE_SERV","")).strip(),
        fmt_br(vl_doc),
        str(r0.get("IND_PGTO","")).strip(),
        fmt_br(vl_desc),
        fmt_br(vl_bc_pis),
        aliq_pis,
        fmt_br(vl_pis),
        fmt_br(vl_bc_cof),
        aliq_cof,
        fmt_br(vl_cof),
        str(r0.get("COD_CTA","")).strip(),
        str(r0.get("COD_MUN","")).strip(),
    ])

    lines = [a100]
    for i, r in enumerate(grupo, start=1):
        a170 = to_pipe(["A170",
            str(i),
            str(r.get("COD_ITEM","")).strip(),
            str(r.get("DESCR_COMPL","")).strip(),
            fmt_valor(r.get("VL_ITEM","")),
            fmt_valor(r.get("VL_DESC_ITEM","")),
            str(r.get("NAT_BC_CRED","")).strip(),
            str(r.get("IND_ORIG_CRED","")).strip(),
            str(r.get("CST_PIS","")).strip(),
            fmt_valor(r.get("VL_BC_PIS_ITEM","")),
            fmt_valor(r.get("ALIQ_PIS_ITEM","")),
            fmt_br(soma_valores(r.get("VL_PIS_ITEM","0"))),
            str(r.get("CST_COFINS","")).strip(),
            fmt_valor(r.get("VL_BC_COFINS_ITEM","")),
            fmt_valor(r.get("ALIQ_COFINS_ITEM","")),
            fmt_br(soma_valores(r.get("VL_COFINS_ITEM","0"))),
            str(r.get("COD_CTA_ITEM","")).strip(),
            str(r.get("COD_CCUS","")).strip(),
        ])
        lines.append(a170)
    return lines

# ── C100/C170 ────────────────────────────────────────────────────────────────
def build_c_block(grupo: list) -> list:
    r0 = grupo[0]
    vl_doc    = soma_valores(*[r.get("VL_ITEM","0") for r in grupo])
    vl_desc   = soma_valores(*[r.get("VL_DESC_ITEM","0") for r in grupo])
    vl_merc   = vl_doc - vl_desc
    vl_pis    = soma_valores(*[r.get("VL_PIS_ITEM","0") for r in grupo])
    vl_cof    = soma_valores(*[r.get("VL_COFINS_ITEM","0") for r in grupo])
    vl_icms   = soma_valores(*[r.get("VL_ICMS_ITEM","0") for r in grupo])
    vl_ipi    = soma_valores(*[r.get("VL_IPI_ITEM","0") for r in grupo])

    c100 = to_pipe(["C100",
        str(r0.get("IND_OPER","")).strip(),
        str(r0.get("IND_EMIT","")).strip(),
        str(r0.get("COD_PART\n(Fornecedor/Cliente)","")).strip(),
        str(r0.get("COD_MOD","")).strip(),
        str(r0.get("COD_SIT","")).strip(),
        str(r0.get("SER","")).strip(),
        str(r0.get("NUM_DOC","")).strip(),
        str(r0.get("CHV_NFE","")).strip(),
        str(r0.get("DT_DOC","")).strip(),
        str(r0.get("DT_E_S","")).strip(),
        fmt_br(vl_doc),
        str(r0.get("IND_PGTO","")).strip(),
        fmt_br(vl_desc),
        fmt_br(soma_valores(r0.get("VL_ABAT_NT","0"))),
        fmt_br(vl_merc),
        str(r0.get("IND_FRT","")).strip(),
        fmt_valor(r0.get("VL_FRT","")),
        fmt_valor(r0.get("VL_SEG","")),
        fmt_valor(r0.get("VL_OUT_DA","")),
        "0,00",          # VL_BC_ICMS – consolidado nos itens
        fmt_br(vl_icms),
        "0,00",          # VL_BC_ICMS_ST
        "0,00",          # VL_ICMS_ST
        fmt_br(vl_ipi),
        fmt_br(vl_pis),
        fmt_br(vl_cof),
        "0,00",          # VL_PIS_ST
        "0,00",          # VL_COFINS_ST
        str(r0.get("COD_CTA","")).strip(),
    ])

    lines = [c100]
    for i, r in enumerate(grupo, start=1):
        c170 = to_pipe(["C170",
            str(i),
            str(r.get("COD_ITEM","")).strip(),
            str(r.get("DESCR_COMPL","")).strip(),
            fmt_valor(r.get("QTD","")),
            str(r.get("UNID","")).strip(),
            fmt_valor(r.get("VL_ITEM","")),
            fmt_valor(r.get("VL_DESC_ITEM","")),
            str(r.get("IND_MOV","")).strip(),
            str(r.get("CST_ICMS","")).strip(),
            str(r.get("CFOP","")).strip(),
            str(r.get("COD_NAT","")).strip(),
            fmt_valor(r.get("VL_BC_ICMS_ITEM","")),
            fmt_valor(r.get("ALIQ_ICMS","")),
            fmt_valor(r.get("VL_ICMS_ITEM","")),
            fmt_valor(r.get("VL_BC_ICMS_ST_ITEM","")),
            fmt_valor(r.get("ALIQ_ST","")),
            fmt_valor(r.get("VL_ICMS_ST_ITEM","")),
            str(r.get("IND_APUR","")).strip() if "IND_APUR" in r else "",
            str(r.get("CST_IPI","")).strip(),
            str(r.get("COD_ENQ","")).strip() if "COD_ENQ" in r else "",
            fmt_valor(r.get("VL_BC_IPI","")),
            fmt_valor(r.get("ALIQ_IPI","")),
            fmt_valor(r.get("VL_IPI_ITEM","")),
            str(r.get("CST_PIS","")).strip(),
            fmt_valor(r.get("VL_BC_PIS_ITEM","")),
            fmt_valor(r.get("ALIQ_PIS_%","")),
            fmt_valor(r.get("QUANT_BC_PIS","")),
            fmt_valor(r.get("ALIQ_PIS_R$","")),
            fmt_valor(r.get("VL_PIS_ITEM","")),
            str(r.get("CST_COFINS","")).strip(),
            fmt_valor(r.get("VL_BC_COFINS_ITEM","")),
            fmt_valor(r.get("ALIQ_COFINS_%","")),
            fmt_valor(r.get("QUANT_BC_COFINS","")),
            fmt_valor(r.get("ALIQ_COFINS_R$","")),
            fmt_valor(r.get("VL_COFINS_ITEM","")),
            str(r.get("COD_CTA_ITEM","")).strip(),
        ])
        lines.append(c170)
    return lines

# ── D100/D101/D105 ───────────────────────────────────────────────────────────
def build_d_block(row: dict) -> list:
    d100 = to_pipe(["D100",
        str(row.get("IND_OPER","")).strip(),
        str(row.get("IND_EMIT","")).strip(),
        str(row.get("COD_PART\n(Transportadora)","")).strip(),
        str(row.get("COD_MOD","")).strip(),
        str(row.get("COD_SIT","")).strip(),
        str(row.get("SER","")).strip(),
        str(row.get("SUB","")).strip(),
        str(row.get("NUM_DOC","")).strip(),
        str(row.get("CHV_CTE","")).strip(),
        str(row.get("DT_DOC","")).strip(),
        str(row.get("DT_A_P","")).strip(),
        str(row.get("TP_CT_e","")).strip(),
        str(row.get("CHV_CTE_REF","")).strip(),
        fmt_valor(row.get("VL_DOC","")),
        fmt_valor(row.get("VL_DESC","")),
        str(row.get("IND_FRT","")).strip(),
        fmt_valor(row.get("VL_SERV","")),
        fmt_valor(row.get("VL_BC_ICMS","")),
        fmt_valor(row.get("VL_ICMS","")),
        fmt_valor(row.get("VL_NT","")),
        str(row.get("COD_CTA","")).strip(),
        str(row.get("COD_INF","")).strip(),
    ])
    d101 = to_pipe(["D101",
        str(row.get("IND_NAT_FRT","")).strip(),
        fmt_valor(row.get("VL_ITEM_PIS","")),
        str(row.get("CST_PIS","")).strip(),
        str(row.get("NAT_BC_CRED","")).strip(),
        fmt_valor(row.get("VL_BC_PIS","")),
        fmt_valor(row.get("ALIQ_PIS_%","")),
        fmt_valor(row.get("QUANT_BC_PIS","")),
        fmt_valor(row.get("ALIQ_PIS_R$","")),
        fmt_valor(row.get("VL_PIS","")),
        str(row.get("COD_CTA_PIS","")).strip(),
        str(row.get("COD_CCUS_PIS","")).strip(),
    ])
    d105 = to_pipe(["D105",
        str(row.get("IND_NAT_FRT","")).strip(),
        fmt_valor(row.get("VL_ITEM_COF","")),
        str(row.get("CST_COFINS","")).strip(),
        str(row.get("NAT_BC_CRED","")).strip(),
        fmt_valor(row.get("VL_BC_COFINS","")),
        fmt_valor(row.get("ALIQ_COFINS_%","")),
        fmt_valor(row.get("QUANT_BC_COF","")),
        fmt_valor(row.get("ALIQ_COFINS_R$","")),
        fmt_valor(row.get("VL_COFINS","")),
        str(row.get("COD_CTA_COF","")).strip(),
        str(row.get("COD_CCUS_COF","")).strip(),
    ])
    return [d100, d101, d105]

# ── F100 / F120 / F130 ───────────────────────────────────────────────────────
def build_f100(row: dict) -> str:
    return to_pipe(["F100",
        str(row.get("IND_OPER","")).strip(),
        str(row.get("COD_PART","")).strip(),
        str(row.get("DT_OPER","")).strip(),
        fmt_valor(row.get("VL_OPER","")),
        str(row.get("COD_ITEM","")).strip(),
        str(row.get("DES_COMPL","")).strip(),
        str(row.get("COD_CTA","")).strip(),
        str(row.get("CST_PIS","")).strip(),
        fmt_valor(row.get("VL_BC_PIS","")),
        fmt_valor(row.get("ALIQ_PIS_%","")),
        fmt_valor(row.get("QUANT_BC_PIS","")),
        fmt_valor(row.get("ALIQ_PIS_R$","")),
        fmt_valor(row.get("VL_PIS","")),
        str(row.get("CST_COFINS","")).strip(),
        fmt_valor(row.get("VL_BC_COFINS","")),
        fmt_valor(row.get("ALIQ_COFINS_%","")),
        fmt_valor(row.get("QUANT_BC_COF","")),
        fmt_valor(row.get("ALIQ_COFINS_R$","")),
        fmt_valor(row.get("VL_COFINS","")),
        str(row.get("NAT_BC_CRED","")).strip(),
        str(row.get("IND_ORIG_CRED","")).strip(),
        str(row.get("COD_CCUS","")).strip(),
        str(row.get("COD_MOD","")).strip(),
        str(row.get("GRUPO_TENSAO","")).strip(),
    ])

def build_f120(row: dict) -> str:
    return to_pipe(["F120",
        str(row.get("NAT_BC_CRED","")).strip(),
        str(row.get("IDENT_BEM_IMOB","")).strip(),
        str(row.get("IND_ORIG_CRED","")).strip(),
        str(row.get("IND_UTIL_BEM_IMOB","")).strip(),
        fmt_valor(row.get("VL_OPER_DEP","")),
        fmt_valor(row.get("PARC_NAO_BC","")),
        str(row.get("CST_PIS","")).strip(),
        fmt_valor(row.get("VL_BC_PIS","")),
        fmt_valor(row.get("ALIQ_PIS_%","")),
        fmt_valor(row.get("QUANT_BC_PIS","")),
        fmt_valor(row.get("ALIQ_PIS_R$","")),
        fmt_valor(row.get("VL_PIS","")),
        str(row.get("CST_COFINS","")).strip(),
        fmt_valor(row.get("VL_BC_COFINS","")),
        fmt_valor(row.get("ALIQ_COFINS_%","")),
        fmt_valor(row.get("QUANT_BC_COF","")),
        fmt_valor(row.get("ALIQ_COFINS_R$","")),
        fmt_valor(row.get("VL_COFINS","")),
        str(row.get("COD_CTA","")).strip(),
        str(row.get("COD_CCUS","")).strip(),
        str(row.get("DES_COMPL","")).strip(),
        str(row.get("COD_MOD","")).strip(),
    ])

def build_f130(row: dict) -> str:
    return to_pipe(["F130",
        str(row.get("NAT_BC_CRED","")).strip(),
        str(row.get("IDENT_BEM_IMOB","")).strip(),
        str(row.get("IND_ORIG_CRED","")).strip(),
        str(row.get("IND_UTIL_BEM_IMOB","")).strip(),
        str(row.get("MES_OPER_AQUIS","")).strip(),
        fmt_valor(row.get("VL_OPER_AQUIS","")),
        fmt_valor(row.get("PARC_NAO_BC","")),
        fmt_valor(row.get("VL_BC_CRED","")),
        str(row.get("IND_NR_PARC","")).strip(),
        str(row.get("CST_PIS","")).strip(),
        fmt_valor(row.get("VL_BC_PIS","")),
        fmt_valor(row.get("ALIQ_PIS_%","")),
        fmt_valor(row.get("QUANT_BC_PIS","")),
        fmt_valor(row.get("ALIQ_PIS_R$","")),
        fmt_valor(row.get("VL_PIS","")),
        str(row.get("CST_COFINS","")).strip(),
        fmt_valor(row.get("VL_BC_COFINS","")),
        fmt_valor(row.get("ALIQ_COFINS_%","")),
        fmt_valor(row.get("QUANT_BC_COF","")),
        fmt_valor(row.get("ALIQ_COFINS_R$","")),
        fmt_valor(row.get("VL_COFINS","")),
        str(row.get("COD_CTA","")).strip(),
        str(row.get("COD_CCUS","")).strip(),
        str(row.get("DES_COMPL","")).strip(),
        str(row.get("COD_MOD","")).strip(),
    ])

# ─────────────────────────────────────────────────────────────────────────────
# Recálculo do 9900 e 9999
# ─────────────────────────────────────────────────────────────────────────────
def recalc_9900(lines: list) -> list:
    """Remove bloco 9 existente e reconstrói 9900 + 9999."""
    # conta registros (excluindo o próprio bloco 9)
    counter = defaultdict(int)
    new_lines = []
    in_bloco9 = False

    for line in lines:
        ls = line.strip()
        if ls.startswith("|9001|") or ls.startswith("|9900|") or \
           ls.startswith("|9990|") or ls.startswith("|9999|"):
            in_bloco9 = True
            continue
        if in_bloco9:
            continue
        new_lines.append(line)
        if ls.startswith("|"):
            reg = ls.strip("|").split("|")[0]
            counter[reg] += 1

    # monta bloco 9
    counter["9001"] = 1
    bloco9 = ["|9001|0|\n"]
    for reg in sorted(counter.keys()):
        bloco9.append(f"|9900|{reg}|{counter[reg]}|\n")
    # conta os próprios 9900
    n9900 = len(bloco9) - 1 + 1  # +1 para o próprio 9900|9900
    bloco9.append(f"|9900|9900|{n9900}|\n")
    bloco9.append(f"|9900|9990|1|\n")
    bloco9.append(f"|9900|9999|1|\n")
    bloco9.append("|9990|1|\n")

    total = sum(counter.values()) + len(bloco9) + 1  # +1 = 9999
    bloco9.append(f"|9999|{total + len(bloco9)}|\n")

    return new_lines + bloco9

# ─────────────────────────────────────────────────────────────────────────────
# Motor principal de retificação
# ─────────────────────────────────────────────────────────────────────────────
def get_col(row: dict, *keys) -> str:
    """Tenta várias chaves alternativas e retorna a primeira encontrada."""
    for k in keys:
        if k in row and str(row[k]).strip():
            return str(row[k]).strip()
    return ""

def retiifica(xlsx_path: str, txt_path: str, out_path: str):
    print(f"\n{'='*60}")
    print(f"Planilha : {xlsx_path}")
    print(f"TXT orig : {txt_path}")
    print(f"TXT saída: {out_path}")
    print('='*60)

    # 1. Lê planilha e TXT
    planilha = load_planilha(xlsx_path)
    lines    = read_txt(txt_path)
    parsed   = parse_txt(lines)

    raw = list(parsed["raw"])   # cópia mutável

    # 2. Marca como retificador no 0000
    for i, line in enumerate(raw):
        if line.strip().startswith("|0000|"):
            f = parse_pipe(line.strip())
            f[2] = "1"   # IND_RET
            raw[i] = to_pipe(f)
            print("✔ 0000 marcado como retificador (IND_RET=1)")
            break

    # 3. ── Bloco 0: novos cadastros ─────────────────────────────────────────
    bloco0_end = parsed["bloco_0_end"]

    # 3a. 0140 – verifica CNPJ_ESTAB da aba 0_EMPRESA
    linhas_0140_novas = []
    empresas = planilha.get("0_EMPRESA", [])
    for emp in empresas:
        cnpj_estab = str(emp.get("CNPJ_ESTAB", emp.get("CNPJ",""))).strip()
        if not cnpj_estab:
            continue
        if cnpj_estab in parsed["cnpj_0140"]:
            print(f"  0140 CNPJ {cnpj_estab} já existe → mantém existente")
        else:
            linha = build_0140(emp)
            linhas_0140_novas.append(linha)
            print(f"  0140 CNPJ {cnpj_estab} NÃO encontrado → criando novo")

    # 3b. Demais registros do bloco 0
    linhas_bloco0_extra = []

    # 0150
    for row in planilha.get("0150_PARTICIPANTES", []):
        linhas_bloco0_extra.append(build_0150(row))

    # 0190
    cod_0190_existentes = set()
    for line in raw:
        if line.strip().startswith("|0190|"):
            f = parse_pipe(line.strip())
            cod_0190_existentes.add(f[1] if len(f) > 1 else "")
    for row in planilha.get("0190_UNIDADES", []):
        unid = str(row.get("UNID","")).strip()
        if unid and unid not in cod_0190_existentes:
            linhas_bloco0_extra.append(build_0190(row))

    # 0200 – centralizado: insere apenas itens novos, não duplica
    for row in planilha.get("0200_ITENS", []):
        cod = str(row.get("COD_ITEM","")).strip()
        if cod and cod not in parsed["cod_0200"]:
            linhas_bloco0_extra.append(build_0200(row))
            parsed["cod_0200"].add(cod)
            print(f"  0200 item {cod} → inserido no bloco 0 centralizado")
        else:
            print(f"  0200 item {cod} já existe → mantém")

    # 0400
    cod_0400_ex = set()
    for line in raw:
        if line.strip().startswith("|0400|"):
            f = parse_pipe(line.strip())
            cod_0400_ex.add(f[1] if len(f) > 1 else "")
    for row in planilha.get("0400_NAT_REC", []):
        cod = str(row.get("COD_NAT_REC","")).strip()
        if cod and cod not in cod_0400_ex:
            linhas_bloco0_extra.append(build_0400(row))

    # 0500
    cod_0500_ex = set()
    for line in raw:
        if line.strip().startswith("|0500|"):
            f = parse_pipe(line.strip())
            cod_0500_ex.add(f[1] if len(f) > 1 else "")
    for row in planilha.get("0500_PLANO_CONTAS", []):
        cod = str(row.get("COD_CTA","")).strip()
        if cod and cod not in cod_0500_ex:
            linhas_bloco0_extra.append(build_0500(row))

    # Injeta no bloco 0 (antes do 0990)
    todas_0_extra = linhas_0140_novas + linhas_bloco0_extra
    if todas_0_extra and bloco0_end is not None:
        raw[bloco0_end:bloco0_end] = todas_0_extra
        print(f"✔ {len(todas_0_extra)} linhas inseridas no bloco 0")

    # Reparse após inserção no bloco 0
    parsed2 = parse_txt(raw)

    # 4. ── Blocos A, C, D, F por estabelecimento ────────────────────────────
    # Mapeia CNPJ → posição do último registro dentro de cada A010/C010/D010/F010
    def find_x010_ranges(lines_list, bloco_prefix):
        """
        Retorna dict {cnpj: idx_ultimo_reg_antes_do_proximo_X010_ou_Xfim}
        para inserção de novos registros ao final de cada sub-bloco do estabelecimento.
        """
        ranges = {}
        x010_positions = []
        x990 = None
        tag010 = f"|{bloco_prefix}010|"
        tag990 = f"|{bloco_prefix}990|"
        for i, line in enumerate(lines_list):
            ls = line.strip()
            if ls.startswith(tag010):
                f = parse_pipe(ls)
                cnpj = f[1] if len(f) > 1 else ""
                x010_positions.append((i, cnpj))
            if ls.startswith(tag990):
                x990 = i
        # calcula o fim de cada sub-bloco
        for k, (pos, cnpj) in enumerate(x010_positions):
            if k + 1 < len(x010_positions):
                fim = x010_positions[k+1][0]
            else:
                fim = x990 if x990 else len(lines_list)
            ranges[cnpj] = fim   # insere antes desse índice
        return ranges

    def inject_by_cnpj(lines_list, bloco_prefix, cnpj_to_lines: dict):
        """Injeta linhas em cada sub-bloco do estabelecimento correto."""
        offset = 0
        ranges = find_x010_ranges(lines_list, bloco_prefix)
        for cnpj, novas in cnpj_to_lines.items():
            if not novas:
                continue
            if cnpj not in ranges:
                print(f"  ⚠ CNPJ {cnpj} não tem {bloco_prefix}010 no TXT – pulando")
                continue
            pos = ranges[cnpj] + offset
            lines_list[pos:pos] = novas
            offset += len(novas)
            print(f"  ✔ Bloco {bloco_prefix}: {len(novas)} linhas inseridas para CNPJ {cnpj}")
        return lines_list

    # ── Agrupa dados da planilha por CNPJ e NUM_DOC ─────────────────────────
    def agrupar_por_doc(sheet_rows, chave_doc="NUM_DOC"):
        """Agrupa linhas por (CNPJ_ESTAB, PERÍODO, NUM_DOC)."""
        grupos = defaultdict(list)
        for row in sheet_rows:
            cnpj   = str(row.get("CNPJ_ESTAB","")).strip()
            periodo= str(row.get("PERÍODO\n(MMAAAA)","")).strip()
            ndoc   = str(row.get(chave_doc,"")).strip()
            if cnpj and ndoc:
                grupos[(cnpj, periodo, ndoc)].append(row)
        return grupos

    # ── Bloco A ──────────────────────────────────────────────────────────────
    a_por_cnpj = defaultdict(list)
    grupos_a = agrupar_por_doc(planilha.get("A100_A170", []))
    for (cnpj, per, ndoc), grupo in grupos_a.items():
        a_por_cnpj[cnpj].extend(build_a_block(grupo))
    raw = inject_by_cnpj(raw, "A", dict(a_por_cnpj))

    # ── Bloco C ──────────────────────────────────────────────────────────────
    c_por_cnpj = defaultdict(list)
    grupos_c = agrupar_por_doc(planilha.get("C100_C170", []))
    for (cnpj, per, ndoc), grupo in grupos_c.items():
        c_por_cnpj[cnpj].extend(build_c_block(grupo))
    raw = inject_by_cnpj(raw, "C", dict(c_por_cnpj))

    # ── Bloco D ──────────────────────────────────────────────────────────────
    d_por_cnpj = defaultdict(list)
    for row in planilha.get("D100_D101_D105", []):
        cnpj = str(row.get("CNPJ_ESTAB","")).strip()
        if cnpj:
            d_por_cnpj[cnpj].extend(build_d_block(row))
    raw = inject_by_cnpj(raw, "D", dict(d_por_cnpj))

    # ── Bloco F ──────────────────────────────────────────────────────────────
    f_por_cnpj = defaultdict(list)
    for row in planilha.get("F100", []):
        cnpj = str(row.get("CNPJ_ESTAB","")).strip()
        if cnpj:
            f_por_cnpj[cnpj].append(build_f100(row))
    for row in planilha.get("F120", []):
        cnpj = str(row.get("CNPJ_ESTAB","")).strip()
        if cnpj:
            f_por_cnpj[cnpj].append(build_f120(row))
    for row in planilha.get("F130", []):
        cnpj = str(row.get("CNPJ_ESTAB","")).strip()
        if cnpj:
            f_por_cnpj[cnpj].append(build_f130(row))
    raw = inject_by_cnpj(raw, "F", dict(f_por_cnpj))

    # 5. Recalcula bloco 9
    raw = recalc_9900(raw)
    print("✔ Bloco 9 recalculado")

    # 6. Salva
    write_txt(out_path, raw)
    print(f"\n✅ Arquivo retificado salvo em: {out_path}")
    print(f"   Total de linhas: {len(raw)}")


# ─────────────────────────────────────────────────────────────────────────────
# Execução
# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Uso: python retifica_efd.py planilha.xlsx EFD_orig.txt EFD_ret.txt")
        sys.exit(1)
    retiifica(sys.argv[1], sys.argv[2], sys.argv[3])
